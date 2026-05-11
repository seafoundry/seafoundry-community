#!/usr/bin/env python3
"""Generate the SeaFoundry Genetics CSV Import Template (XLSX).

Creates an Excel workbook with:
- A data entry sheet with dropdown validation for enumerated fields
- A reference sheet documenting every column, its rules, and examples
- A hidden lookup sheet backing the dropdown lists

Run:
    python3 scripts/generate_genetics_template.py

Output:
    templates/SeaFoundry_Genetics_Import_Template.xlsx

Requires:
    pip install openpyxl
"""

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Constants ────────────────────────────────────────────────────────────────
# Species codes must match lib/models/species.dart Species.coralSpecies.

SPECIES_CODES = [
    ("ACER", "Acropora cervicornis"),
    ("APAL", "Acropora palmata"),
    ("OFAV", "Orbicella faveolata"),
    ("PAST", "Porites astreoides"),
    ("PPOR", "Porites porites"),
    ("MCAV", "Montastraea cavernosa"),
    ("DLAB", "Diploria labyrinthiformis"),
    ("SSID", "Siderastrea siderea"),
    ("PSTR", "Pseudodiploria strigosa"),
    ("CNAT", "Colpophyllia natans"),
    ("DCYL", "Dendrogyra cylindrus"),
]

PROVENANCE_TYPES = [
    ("wild", "Wild Collection -- founder/broodstock genets collected from reef"),
    ("cohort", "Nursery Cohort -- nursery-reared cohort (asexual propagation batch)"),
    ("graduatedIndividual", "Graduated Individual -- promoted from cohort to own genet"),
    ("transfer", "Transfer / Import -- received from another organization"),
    ("unknown", "Unknown -- provenance not determined"),
]

# Physical form IDs for coral (from lib/services/physical_form_data.dart)
PHYSICAL_FORMS = [
    ("fragment", "Fragment (Clipping)"),
    ("colony", "Colony"),
    ("spawning_colony", "Spawning Colony (Broodstock)"),
    ("mounted_individual", "Mounted Individual"),
    ("shared_substrate", "Shared Substrate"),
    ("settlement_substrate", "Settlement Substrate"),
    ("liquid_suspension", "Liquid Suspension"),
]

ORGANISM_KINDS = ["coral"]

BOOLEAN_VALUES = ["false", "true"]

# Column definitions: (header, width, required, description, example, rules)
# Must match _geneticsColumns in lib/constants/csv_schema.dart plus
# organism-record fields needed by genetics_csv_importer.dart.
COLUMNS = [
    (
        "provenanceId",
        22,
        True,
        "Unique lineage identifier for the genet (PID format)",
        "PID-ACER-0001",
        "Format: PID-XXXX-NNNN where XXXX is 2-4 letter species slug "
        "and NNNN is a zero-padded counter. Must start with 'PID-'.",
    ),
    (
        "localId",
        18,
        True,
        "Local identifier for the genet (e.g. ACER-001). "
        "Shared by all records in the same genet.",
        "ACER-001",
        "Free text, typically SPECIES-SEQ. Used as the organism's "
        "localId if no separate organism localId is provided.",
    ),
    (
        "name",
        20,
        True,
        "Display name for the genet",
        "Staghorn Alpha",
        "3-30 characters. Letters, numbers, and spaces only. "
        "No special characters.",
    ),
    (
        "speciesId",
        14,
        True,
        "Species code (use dropdown)",
        "ACER",
        "Must be one of the supported species codes. "
        "Case-insensitive on import.",
    ),
    (
        "organismKind",
        14,
        True,
        "Organism type (always 'coral' for community tier)",
        "coral",
        "Only 'coral' is supported. Defaults to 'coral' if blank.",
    ),
    (
        "provenanceType",
        22,
        False,
        "Provenance type (use dropdown). Controls default life stage.",
        "wild",
        "wild | cohort | graduatedIndividual | transfer | unknown. "
        "Defaults to 'wild' if blank.",
    ),
    (
        "clonalId",
        18,
        False,
        "Clonal identifier from a naming system (CRC, Mote, CRF, etc.)",
        "AC-1",
        "Free text. Used to cross-reference genets across organizations. "
        "No format restriction.",
    ),
    (
        "accessionNumber",
        22,
        False,
        "Accession number from a registry or collection catalog",
        "ACER-2019-001",
        "Free text. Typically follows the pattern SPECIES-YEAR-SEQ "
        "but no strict format is enforced.",
    ),
    (
        "aliases",
        30,
        False,
        "Alternative identifiers from other organizations",
        "K2; NOVA-ACER-001; AC-1-Mote",
        "Semicolon-separated list of alias values. Each value becomes "
        "an alias record. For structured aliases use JSON: "
        '[{"sourceSystem":"CRF","value":"K2"}]',
    ),
    (
        "notes",
        24,
        False,
        "Genet-level notes (distinct from provenance.notes)",
        "Robust growth, high survival",
        "Free text. General notes about this genet.",
    ),
    (
        "parentGameteIds",
        22,
        False,
        "Parent gamete provenance IDs (for cohort genets)",
        "PID-ACER-0010; PID-ACER-0020",
        "Semicolon-separated list of parent PID values. "
        "Only relevant for cohort provenance type.",
    ),
    (
        "donorGenotypeId",
        20,
        False,
        "Donor genotype ID (for graduated individuals)",
        "PID-ACER-0005",
        "PID of the donor genet. Only relevant for "
        "graduatedIndividual provenance type.",
    ),
    (
        "archived",
        12,
        False,
        "Whether this genet is archived (use dropdown)",
        "false",
        "true/false, yes/no, 1/0 accepted. Blank = no change.",
    ),
    (
        "provenance.habitatType",
        22,
        False,
        "Habitat type where the genet was collected",
        "patch reef",
        "Free text describing the collection habitat.",
    ),
    (
        "provenance.collectionDate",
        24,
        False,
        "Date the genet was collected from the wild",
        "2019-06-15",
        "ISO-8601 date format (YYYY-MM-DD). Only relevant for "
        "wild provenance type.",
    ),
    (
        "provenance.notes",
        30,
        False,
        "Additional provenance-specific metadata notes",
        "Collected at Looe Key, 5m depth",
        "Free text. Provenance metadata only -- for genet-level "
        "notes use the 'notes' column instead.",
    ),
    # ── Organism Record columns (required by genetics importer) ──
    (
        "groupId",
        28,
        True,
        "URL path of the parent group/structure in SeaFoundry",
        "organizations/demo-org/sites/nursery-1/groups/tree-1",
        "Must match an existing group's urlPath. "
        "Find this in the app URL bar when viewing the group.",
    ),
    (
        "physicalFormId",
        20,
        True,
        "Physical form of the coral (use dropdown)",
        "fragment",
        "fragment | colony | spawning_colony | mounted_individual | "
        "shared_substrate | settlement_substrate | liquid_suspension",
    ),
    (
        "Record Name",
        20,
        False,
        "User-friendly name for this organism record",
        "Fluffy",
        "3-50 characters. Letters, numbers, spaces, hyphens, underscores. "
        "Falls back to localId if blank.",
    ),
    (
        "Quantity",
        12,
        False,
        "Number of individuals in this holding",
        "1",
        "Whole number, 1-10000. Defaults to 1 if blank.",
    ),
    (
        "Size",
        10,
        False,
        "Measured dimension in centimeters",
        "5.2",
        "Decimal number, 0.1-1000 cm. Leave blank if not measured.",
    ),
]


# ── Styles ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B5E8C", end_color="1B5E8C", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
EXAMPLE_FONT = Font(name="Calibri", italic=True, size=10, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1B5E8C")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
REF_HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
REF_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")


def _build_lookup_sheet(wb):
    """Hidden sheet with list values for data validation."""
    ws = wb.create_sheet("_Lookups")

    # Species codes (column A)
    ws["A1"] = "speciesId"
    ws["A1"].font = Font(bold=True)
    for i, (code, _) in enumerate(SPECIES_CODES, start=2):
        ws.cell(row=i, column=1, value=code)

    # Provenance types (column B)
    ws["B1"] = "provenanceType"
    ws["B1"].font = Font(bold=True)
    for i, (pt, _) in enumerate(PROVENANCE_TYPES, start=2):
        ws.cell(row=i, column=2, value=pt)

    # Organism kinds (column C)
    ws["C1"] = "organismKind"
    ws["C1"].font = Font(bold=True)
    for i, ok in enumerate(ORGANISM_KINDS, start=2):
        ws.cell(row=i, column=3, value=ok)

    # Boolean values (column D)
    ws["D1"] = "boolean"
    ws["D1"].font = Font(bold=True)
    for i, bv in enumerate(BOOLEAN_VALUES, start=2):
        ws.cell(row=i, column=4, value=bv)

    # Physical forms (column E)
    ws["E1"] = "physicalFormId"
    ws["E1"].font = Font(bold=True)
    for i, (fid, _) in enumerate(PHYSICAL_FORMS, start=2):
        ws.cell(row=i, column=5, value=fid)

    ws.sheet_state = "hidden"
    return ws


def _build_data_sheet(wb):
    """Main data entry sheet with headers, validation, and comment-prefixed helper rows."""
    ws = wb.create_sheet("Genetics Import", 0)

    all_columns = list(COLUMNS)

    # ── Metadata rows (rows 1-2) ────────────────────────────────────────
    # Key in column A, value in column B (CSV parser expects 2-cell rows).
    meta_font = Font(name="Calibri", size=9, color="999999")
    ws.cell(row=1, column=1, value="provenanceCsvVersion").font = meta_font
    ws.cell(row=1, column=2, value="2025.10").font = meta_font
    ws.cell(row=2, column=1, value="provenanceCsvTemplate").font = meta_font
    ws.cell(row=2, column=2, value="genetics").font = meta_font

    # ── Column headers (row 3) ───────────────────────────────────────────
    header_row = 3
    for col_idx, (header, width, required, desc, _, _) in enumerate(
        all_columns, start=1
    ):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Comment-prefixed helper rows (rows 4-5) ─────────────────────────
    # The CSV parser skips lines starting with '#', so these won't be
    # imported as data.  They serve as in-sheet documentation.

    # Row 4: Required / Optional indicator (prefixed with # in col A)
    for col_idx, (_, _, required, _, _, _) in enumerate(all_columns, start=1):
        label = "REQUIRED" if required else "optional"
        if col_idx == 1:
            label = "# " + label  # Parser skip prefix
        cell = ws.cell(row=header_row + 1, column=col_idx, value=label)
        cell.font = Font(
            name="Calibri",
            bold=required,
            size=9,
            color="CC0000" if required else "666666",
        )
        cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Row 5: Example values (prefixed with # in col A)
    for col_idx, (_, _, _, _, example, _) in enumerate(all_columns, start=1):
        value = example
        if col_idx == 1:
            value = "# " + (example or "")  # Parser skip prefix
        cell = ws.cell(row=header_row + 2, column=col_idx, value=value)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    # ── Data validation ──────────────────────────────────────────────────
    data_start_row = header_row + 3  # Row 6
    data_end_row = 1000

    def _col_index(col_name):
        return next(
            i for i, (h, *_) in enumerate(all_columns, start=1) if h == col_name
        )

    def _add_list_validation(col_name, formula, allow_blank, error_title, error_msg,
                              prompt_title, prompt_msg):
        col = _col_index(col_name)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=allow_blank,
            showErrorMessage=True,
            errorTitle=error_title,
            error=error_msg,
            promptTitle=prompt_title,
            prompt=prompt_msg,
            showInputMessage=True,
        )
        dv.add(
            f"{get_column_letter(col)}{data_start_row}:"
            f"{get_column_letter(col)}{data_end_row}"
        )
        ws.add_data_validation(dv)

    # speciesId dropdown
    _add_list_validation(
        "speciesId",
        f"_Lookups!$A$2:$A${len(SPECIES_CODES) + 1}",
        allow_blank=False,
        error_title="Invalid Species",
        error_msg="Select a valid species code from the dropdown.",
        prompt_title="Species Code",
        prompt_msg=(
            "Select the coral species code.\n\n"
            "ACER = A. cervicornis\n"
            "APAL = A. palmata\n"
            "OFAV = O. faveolata\n"
            "DCYL = D. cylindrus\netc."
        ),
    )

    # organismKind dropdown
    _add_list_validation(
        "organismKind",
        "_Lookups!$C$2:$C$2",
        allow_blank=True,
        error_title="Invalid Organism Kind",
        error_msg="Only 'coral' is supported.",
        prompt_title="Organism Kind",
        prompt_msg="Community tier only supports 'coral'.",
    )

    # provenanceType dropdown
    _add_list_validation(
        "provenanceType",
        f"_Lookups!$B$2:$B${len(PROVENANCE_TYPES) + 1}",
        allow_blank=True,
        error_title="Invalid Provenance Type",
        error_msg="Select a valid provenance type from the dropdown.",
        prompt_title="Provenance Type",
        prompt_msg=(
            "wild = founder/broodstock\n"
            "cohort = asexual propagation batch\n"
            "graduatedIndividual = promoted from cohort\n"
            "transfer = received from another org\n"
            "unknown = not determined"
        ),
    )

    # archived dropdown
    _add_list_validation(
        "archived",
        "_Lookups!$D$2:$D$3",
        allow_blank=True,
        error_title="Invalid Value",
        error_msg="Use 'true' or 'false'.",
        prompt_title="Archived",
        prompt_msg="Set to 'true' to archive the genet, 'false' or blank to keep active.",
    )

    # physicalFormId dropdown
    _add_list_validation(
        "physicalFormId",
        f"_Lookups!$E$2:$E${len(PHYSICAL_FORMS) + 1}",
        allow_blank=False,
        error_title="Invalid Physical Form",
        error_msg="Select a valid physical form from the dropdown.",
        prompt_title="Physical Form",
        prompt_msg=(
            "fragment = Fragment/Clipping\n"
            "colony = Colony\n"
            "spawning_colony = Broodstock\n"
            "mounted_individual = Mounted\n"
            "shared_substrate = Shared Substrate"
        ),
    )

    # Name length validation
    name_col = _col_index("name")
    dv_name = DataValidation(
        type="textLength",
        operator="between",
        formula1="3",
        formula2="30",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Genet Name",
        error="Genet name must be 3-30 characters. Letters, numbers, and spaces only.",
        promptTitle="Genet Name",
        prompt="Enter a display name for this genet.\n3-30 characters.\nLetters, numbers, and spaces only.",
        showInputMessage=True,
    )
    dv_name.add(
        f"{get_column_letter(name_col)}{data_start_row}:"
        f"{get_column_letter(name_col)}{data_end_row}"
    )
    ws.add_data_validation(dv_name)

    # Freeze panes below headers + helper rows
    ws.freeze_panes = f"A{data_start_row}"

    return ws, all_columns


def _build_reference_sheet(wb, all_columns):
    """Reference sheet documenting field rules, species codes, and import behavior."""
    ws = wb.create_sheet("Reference Guide")

    row = 1

    # ── Title ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(
        row=row, column=1,
        value="SeaFoundry Genetics Import Template - Reference Guide",
    )
    cell.font = Font(name="Calibri", bold=True, size=14, color="1B5E8C")
    row += 2

    # ── How to Use ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="HOW TO USE THIS TEMPLATE").font = SECTION_FONT
    row += 1

    instructions = [
        "1. Fill in genet data on the 'Genetics Import' sheet starting at row 6.",
        "2. Row 3 contains column headers -- do not modify these.",
        "3. Row 4 (REQUIRED/optional) and row 5 (examples) start with # and are "
        "skipped by the importer. You may leave, edit, or delete them.",
        "4. Use the dropdown arrows for species, provenance type, organism kind, "
        "physical form, and archived fields.",
        "5. Rows 1-2 contain CSV metadata (version & template) -- do not modify.",
        "6. Save as CSV (comma delimited) via File > Save As, then use "
        "SeaFoundry's Import > Genetics to upload.",
        "7. The import wizard validates all rows before committing. Fix any errors "
        "and re-upload.",
        "",
        "NOTE: The import only accepts .csv files. This XLSX template provides "
        "dropdown validation for data entry. Save as CSV before uploading.",
    ]
    for line in instructions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name="Calibri", size=10)
        c.alignment = WRAP_ALIGNMENT
        row += 1
    row += 1

    # ── Column Reference Table ───────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="COLUMN REFERENCE").font = SECTION_FONT
    row += 1

    ref_headers = [
        "Column Name", "Required", "Description", "Example", "Validation Rules",
    ]
    for col_idx, header in enumerate(ref_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = REF_HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for header, _, required, desc, example, rules in all_columns:
        ws.cell(row=row, column=1, value=header).border = THIN_BORDER
        req_cell = ws.cell(
            row=row, column=2, value="Yes" if required else "No",
        )
        req_cell.border = THIN_BORDER
        req_cell.font = Font(
            name="Calibri", bold=required,
            color="CC0000" if required else "333333",
        )
        ws.cell(row=row, column=3, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=4, value=example).border = THIN_BORDER
        ws.cell(row=row, column=5, value=rules).border = THIN_BORDER
        for c in range(1, 6):
            ws.cell(row=row, column=c).alignment = WRAP_ALIGNMENT
        row += 1
    row += 1

    # ── Species Code Reference ───────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="SPECIES CODES").font = SECTION_FONT
    row += 1

    for col_idx, header in enumerate(["Code", "Scientific Name"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = REF_HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.border = THIN_BORDER
    row += 1
    for code, name in SPECIES_CODES:
        ws.cell(row=row, column=1, value=code).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        row += 1
    row += 1

    # ── Provenance Types Reference ───────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="PROVENANCE TYPES").font = SECTION_FONT
    row += 1

    for col_idx, header in enumerate(["Value", "Description"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = REF_HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.border = THIN_BORDER
    row += 1
    for value, desc in PROVENANCE_TYPES:
        ws.cell(row=row, column=1, value=value).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        row += 1
    row += 1

    # ── Physical Forms Reference ─────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="PHYSICAL FORMS (CORAL)").font = SECTION_FONT
    row += 1

    for col_idx, header in enumerate(["ID", "Display Name"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = REF_HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.border = THIN_BORDER
    row += 1
    for fid, fname in PHYSICAL_FORMS:
        ws.cell(row=row, column=1, value=fid).border = THIN_BORDER
        ws.cell(row=row, column=2, value=fname).border = THIN_BORDER
        row += 1
    row += 1

    # ── Alias Format Reference ───────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="ALIAS FORMAT").font = SECTION_FONT
    row += 1

    alias_notes = [
        "Simple format: Semicolon-separated values   e.g.  K2; NOVA-001; AC-1-Mote",
        "Each value becomes an alias with sourceSystem='custom'.",
        "",
        "JSON format (for structured aliases with source tracking):",
        '[{"sourceSystem": "CRF", "value": "K2"}, '
        '{"sourceSystem": "Mote", "value": "AC-1"}]',
        "",
        "Fields: sourceSystem (optional, default: custom), value (required),",
        "        label (optional, display name), isPrimary (optional, boolean)",
    ]
    for line in alias_notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=line).font = Font(name="Calibri", size=10)
        row += 1
    row += 1

    # ── Provenance ID Format ─────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(
        row=row, column=1, value="PROVENANCE ID (PID) FORMAT",
    ).font = SECTION_FONT
    row += 1

    pid_notes = [
        "Format: PID-XXXX-NNNN",
        "  PID-    Fixed prefix (required)",
        "  XXXX    Species slug, 2-4 uppercase letters (e.g. ACER, AP)",
        "  NNNN    Zero-padded sequential counter (4+ digits)",
        "",
        "Valid examples:   PID-ACER-0001   PID-AP-0042   PID-DCYL-00100",
        "Invalid examples: SF-ACER-0001 (wrong prefix)   0001 (missing prefix)",
        "",
        "IMPORTANT: Do not confuse provenance IDs (PID-...) with Firestore "
        "document IDs. The provenanceId column expects ONLY lineage IDs.",
    ]
    for line in pid_notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=line).font = Font(name="Calibri", size=10)
        row += 1
    row += 1

    # ── Import Behavior Notes ────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="IMPORT BEHAVIOR NOTES").font = SECTION_FONT
    row += 1

    behavior_notes = [
        "- Each row creates BOTH a Genet record AND an OrganismRecord (holding) "
        "linked to it.",
        "- If a genet with the same provenanceId already exists, the existing "
        "genet is reused (no duplicate created).",
        "- If a genet with the same name already exists, the existing genet is "
        "reused.",
        "- New genets are assigned a PID by the system if the provided PID is "
        "not found in the organization.",
        "- clonalId, accessionNumber, and aliases are only set when creating "
        "NEW genets.",
        "- The 'archived' field can update existing genets: set to 'true' to "
        "archive, 'false' to restore.",
        "- groupId must be a valid urlPath of an existing group. Find it in the "
        "browser URL bar when viewing a group.",
        "- physicalFormId is required. Use the dropdown to select a valid coral "
        "physical form.",
        "- localId is required and becomes the genet's local identifier "
        "(e.g. ACER-001).",
        "- Quantity defaults to 1 if blank. Record Name falls back to localId "
        "if blank.",
        "- Default life stage is set by provenance type: wild -> broodstock, "
        "cohort -> juvenile, others -> adult.",
        "- Rows with validation errors are skipped; other rows continue "
        "importing.",
        "- The import wizard shows a preview with error details before "
        "committing.",
    ]
    for line in behavior_notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=line).font = Font(name="Calibri", size=10)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 58

    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    _build_lookup_sheet(wb)
    data_ws, all_columns = _build_data_sheet(wb)
    _build_reference_sheet(wb, all_columns)

    output_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "templates",
    )
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(
        output_dir, "SeaFoundry_Genetics_Import_Template.xlsx",
    )
    wb.save(output_path)
    print(f"Template saved to: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Columns: {len(all_columns)}")
    print(f"  Species codes: {len(SPECIES_CODES)}")
    print(f"  Provenance types: {len(PROVENANCE_TYPES)}")
    print(f"  Physical forms: {len(PHYSICAL_FORMS)}")


if __name__ == "__main__":
    main()
